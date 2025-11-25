#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl

# 打开Excel文件
file_path = './xiaohongshu_content.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# 添加测试标题
# 先检查当前有多少行数据
current_row = sheet.max_row + 1

# 测试标题列表
test_titles = [
    "夏日必备防晒技巧",
    "如何提高工作效率",
    "健康饮食小常识",
    "旅行拍照技巧分享",
    "学习Python的好处"
]

# 添加测试标题到Excel
for title in test_titles:
    sheet.cell(row=current_row, column=1).value = title
    current_row += 1

# 保存并关闭文件
wb.save(file_path)
wb.close()

print(f"已添加{len(test_titles)}个测试标题到{file_path}")
