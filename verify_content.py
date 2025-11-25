#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl

# 打开Excel文件
file_path = './test_xiaohongshu_content.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# 获取最大行数
max_row = sheet.max_row

print("验证生成的内容：")
print("=" * 50)

# 从第2行开始验证
for row in range(2, max_row + 1):
    title = sheet.cell(row=row, column=1).value
    content = sheet.cell(row=row, column=2).value
    
    if title and content:
        length = len(content)
        status = "✅" if length <= 70 else "❌"
        print(f"标题：{title}")
        print(f"内容：{content}")
        print(f"长度：{length} 字符 {status}")
        print("-" * 50)

# 关闭工作簿
wb.close()
