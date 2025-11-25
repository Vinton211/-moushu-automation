#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
小红书PC端自动化脚本主程序
"""

from browser_manager import BrowserManager
from popup_handler import PopupHandler
from login_manager import LoginManager
from publisher import Publisher
from content_reader import ContentReader
import argparse
import traceback
import openpyxl
from openai import OpenAI
import os

def generate_content(title, api_key=None):
    """
    使用DeepSeek API实时生成内容
    
    Args:
        title: 标题
        api_key: DeepSeek API密钥，可选，默认从环境变量获取
    
    Returns:
        str: 生成的内容，限制在70字符以内
    """
    try:
        # 初始化DeepSeek客户端
        client = OpenAI(
            api_key=api_key,
            base_url="https://api.deepseek.com"
        )
        
        # 调用DeepSeek API生成内容
        response = client.chat.completions.create(
            model="deepseek-reasoner",
            messages=[
                {"role": "system", "content": "你是一个小红书内容创作者，请根据给定标题生成一条简洁的小红书内容，不超过700个字符。内容要完整可读"},
                {"role": "user", "content": f"标题：{title}"}
            ],
            max_tokens=1000,
            temperature=1.3
        )
        
        # 获取生成的内容
        content = response.choices[0].message.content.strip()
        
        # 确保内容不超过70字符
        if len(content) > 700:
            content = content[:700]
        
        return content
    except Exception as e:
        print(f"生成内容失败：{e}")
        return f"这是关于{title}的内容，简洁明了，适合快速阅读。"

def process_excel(file_path, api_key=None):
    """
    处理Excel文件，生成内容
    
    Args:
        file_path: Excel文件路径
        api_key: DeepSeek API密钥，可选，默认从环境变量获取
    """
    # 打开Excel文件
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # 获取最大行数
    max_row = sheet.max_row
    
    # 从第2行开始处理（第1行是表头）
    for row in range(2, max_row + 1):
        title = sheet.cell(row=row, column=1).value
        content = sheet.cell(row=row, column=2).value
        
        # 如果标题存在且内容不存在或为空，则生成内容
        if title and not content:
            print(f"处理标题：{title}")
            generated_content = generate_content(title, api_key)
            if generated_content:
                sheet.cell(row=row, column=2).value = generated_content
                print(f"生成内容：{generated_content}")
    
    # 保存文件
    wb.save(file_path)
    wb.close()
    print(f"处理完成，已保存到{file_path}")
class XiaoHongShuPCAutomation:
    """小红书PC自动化主类"""
    
    def __init__(self, headless=False, driver_path='./chromedriver.exe', reuse_browser=False):
        """
        初始化小红书PC自动化对象
        
        Args:
            headless: 是否使用无头模式
            driver_path: ChromeDriver的路径
            reuse_browser: 是否尝试复用已打开的浏览器实例
        """
        # 初始化各个模块
        # 初始化浏览器管理器，负责启动/关闭浏览器并返回 driver 与等待对象
        self.browser_manager = BrowserManager(headless=headless, driver_path=driver_path, reuse_browser=reuse_browser)
        self.driver = None  # 初始化浏览器驱动，后续由 BrowserManager 赋值
        self.wait = None  # 初始化等待对象，后续由 BrowserManager 赋值
        self.popup_handler = None  # 初始化弹窗处理器，负责处理登录弹窗等
        self.login_manager = None  # 初始化登录管理器，负责处理登录流程
        self.publisher = None  # 初始化发布器，负责发布笔记
        self.content_reader = None  # 初始化内容读取器，负责从Excel读取笔记数据
    
    def initialize(self):
        """初始化所有组件"""
        # 初始化浏览器
        # 调用 BrowserManager 初始化浏览器并返回 driver 与等待对象
        self.driver, self.wait = self.browser_manager.initialize_browser()
          
        # 初始化其他组件
        self.popup_handler = PopupHandler(self.driver, self.wait)   
        self.login_manager = LoginManager(self.driver, self.wait, self.popup_handler)
        self.publisher = Publisher(self.driver, self.wait, self.popup_handler)
    
    def open_xiaohongshu(self, is_creator=False):
        """
        打开小红书网站
        
        Args:
            is_creator: 是否打开创作服务平台
        """
        self.login_manager.open_xiaohongshu(is_creator)
    
    def login(self, phone_number=None, password=None, is_creator=False):
        """
        登录小红书
        
        Args:
            phone_number: 手机号（可选，若不提供则使用默认值）
            password: 密码（可选，若不提供则使用验证码登录）
            is_creator: 是否是创作服务平台登录
        
        Returns:
            bool: 登录是否成功
        """
        return self.login_manager.login(phone_number, password, is_creator)
    
    def publish_note(self, note_data):
        """
        发布笔记
        
        Args:
            note_data: 笔记数据，包含标题、内容、图片路径等
        
        Returns:
            bool: 发布是否成功
        """
        return self.publisher.publish_note(note_data)
    
    def read_notes_from_excel(self, excel_file_path):
        """
        从Excel文件中读取笔记数据
        
        Args:
            excel_file_path: Excel文件路径
        
        Returns:
            list: 笔记数据列表
        """
        self.content_reader = ContentReader(excel_file_path)
        try:
            return self.content_reader.read_all_notes()
        finally:
            if self.content_reader:
                self.content_reader.close_workbook()
    
    def close(self):
        """关闭浏览器"""
        self.browser_manager.close_browser()

# 示例用法
if __name__ == "__main__":

    
    
    try:
        # 添加命令行参数
        parser = argparse.ArgumentParser(description='小红书PC自动化脚本')
        parser.add_argument('--keep-browser-open', action='store_true', 
                           help='保持浏览器打开状态，不自动关闭')
        parser.add_argument('--reuse-browser', action='store_true',
                           help='尝试复用已打开的浏览器实例，避免重新登录')
        args = parser.parse_args()
        
        print("=== 小红书PC自动化脚本 ===")
        print("本脚本用于自动化生成Excel内容并发布小红书笔记")
        print("请确保已安装Chrome浏览器和对应版本的ChromeDriver")
        print("ChromeDriver应放置在脚本同一目录下")
        print("\n")
        
        # 1. 处理Excel内容生成
        excel_file = "xiaohongshu_content.xlsx"  # Excel文件路径
        print(f"\n=== 1. 开始处理Excel内容生成 ===")
        print(f"Excel文件路径: {excel_file}")
        
        # 设置DeepSeek API密钥
        api_key = 'sk-6fdf451e9b3f4f4baa5ed7828e583a83'
        
        # 处理Excel文件，生成内容
        process_excel(excel_file, api_key)
        
        # 2. 初始化浏览器并发布笔记
        print(f"\n=== 2. 开始小红书自动发布流程 ===")
        
        # 初始化自动化对象
        print("正在初始化浏览器...")
        xhs_automation = XiaoHongShuPCAutomation(headless=False, reuse_browser=args.reuse_browser)
        xhs_automation.initialize()
        
        # 打开小红书
        print("\n正在打开小红书创作服务平台...")
        xhs_automation.open_xiaohongshu(is_creator=True)
        
        # 检查是否已登录，如果复用浏览器且已登录则跳过登录流程
        login_success = True
        if args.reuse_browser:
            print(f"\n正在检查登录状态...")
            login_success = xhs_automation.login_manager._check_login_status()
            if login_success:
                print("已登录，跳过登录流程")
        
        # 如果未登录，则执行登录流程
        if not login_success:
            print(f"\n正在使用默认手机号登录...")
            print("请确保手机能够接收到验证码")
            login_success = xhs_automation.login(is_creator=True)
        
        if login_success:
            print("\n登录成功！")
            # 从Excel读取笔记数据
            print(f"\n正在从Excel文件 {excel_file} 读取笔记数据...")
            notes = xhs_automation.read_notes_from_excel(excel_file)
            
            if notes:
                print(f"共读取到 {len(notes)} 条笔记数据")
                # 发布第一条笔记
                print("\n正在发布第一条笔记...")
                note_data = notes[0]
                print(f"标题: {note_data['title']}")
                print(f"内容: {note_data['content'][:50]}...")  # 只显示前50个字符
                
                publish_success = xhs_automation.publish_note(note_data)
                
                if publish_success:
                    print("\n笔记发布成功！")
                else:
                    print("\n笔记发布失败！")
            else:
                print("\n没有读取到笔记数据！")
        else:
            print("\n登录失败！")
        
        # 根据命令行参数决定是否关闭浏览器
        if args.keep_browser_open:
            print("\n浏览器将保持打开状态，您可以继续手动操作")
            print("要关闭浏览器，请手动关闭窗口或重新运行脚本不带 --keep-browser-open 参数")
        else:
            # 关闭浏览器
            print("\n正在关闭浏览器...")
            xhs_automation.close()
        
        print("\n=== 脚本执行完毕 ===")
    except Exception as e:
        print(f"\n程序执行过程中发生错误: {e}")

        traceback.print_exc()
