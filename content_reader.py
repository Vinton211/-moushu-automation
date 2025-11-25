#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
内容读取器，负责从Excel文件中读取内容
"""

import openpyxl

class ContentReader:
    """内容读取器类"""
    
    def __init__(self, file_path):
        """
        初始化内容读取器
        
        Args:
            file_path: Excel文件路径
        """
        self.file_path = file_path
        self.workbook = None
        self.sheet = None
    
    def open_workbook(self):
        """打开Excel工作簿"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            # 获取第一个工作表
            self.sheet = self.workbook.active
            print(f"已打开Excel文件: {self.file_path}")
            print(f"当前工作表: {self.sheet.title}")
        except Exception as e:
            print(f"打开Excel文件失败: {e}")
            raise
    
    def read_note_data(self, row_num):
        """
        读取指定行的笔记数据
        
        Args:
            row_num: 行号（从1开始）
        
        Returns:
            dict: 笔记数据
        """
        try:
            if not self.sheet:
                self.open_workbook()
            
            # 假设Excel列的映射关系
            # A列: 标题
            # B列: 正文内容
            # C列: 图片路径（用分号分隔多个图片）
            # D列: 话题标签（用分号分隔多个标签）
            # E列: 分类
            
            note_data = {
                'title': self.sheet.cell(row=row_num, column=1).value or '',
                'content': self.sheet.cell(row=row_num, column=2).value or '',
                'image_paths': [],
                'tags': [],
                'category': self.sheet.cell(row=row_num, column=5).value or ''
            }
            
            # 处理图片路径
            image_paths_str = self.sheet.cell(row=row_num, column=3).value
            if image_paths_str:
                note_data['image_paths'] = [path.strip() for path in image_paths_str.split(';') if path.strip()]
            
            # 处理话题标签
            tags_str = self.sheet.cell(row=row_num, column=4).value
            if tags_str:
                note_data['tags'] = [tag.strip() for tag in tags_str.split(';') if tag.strip()]
            
            return note_data
        except Exception as e:
            print(f"读取笔记数据失败: {e}")
            return {}
    
    def read_all_notes(self):
        """
        读取所有笔记数据
        
        Returns:
            list: 所有笔记数据的列表
        """
        try:
            if not self.sheet:
                self.open_workbook()
            
            notes = []
            # 从第2行开始读取（假设第1行是表头）
            for row in range(2, self.sheet.max_row + 1):
                note_data = self.read_note_data(row)
                if note_data:
                    notes.append(note_data)
            
            print(f"共读取到 {len(notes)} 条笔记数据")
            return notes
        except Exception as e:
            print(f"读取所有笔记数据失败: {e}")
            return []
    
    def close_workbook(self):
        """关闭Excel工作簿"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.sheet = None
            print("已关闭Excel工作簿")
    
    def __enter__(self):
        """上下文管理器进入方法"""
        self.open_workbook()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器退出方法"""
        self.close_workbook()
